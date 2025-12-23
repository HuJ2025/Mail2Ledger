"""
json_to_xlsx_bytes(data) -> BytesIO

Clean export:
- No cell highlights on Description / Dates / Status / Type.
- Smart numeric coercion: strings that are numbers become numbers; others stay text.
- Safer openpyxl values (lists/dicts -> JSON strings) with XML-unsafe chars stripped.
- Nicer sheet polish: frozen panes, tidy widths, subtle header fill, optional zebra rows.
"""
from __future__ import annotations
import json
import re
import math
from datetime import datetime
from io import BytesIO
from typing import Any, Iterable, Tuple, Union, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # strips XML-illegal control chars

# ---- visual constants ----
TITLE_FONT   = Font(bold=True, size=14)
HDR_FONT     = Font(bold=True, size=11)
DATA_FONT    = Font(size=11)
PERCENT_FMT  = "0.00%"

HDR_FILL     = PatternFill("solid", "F2F2F2", "F2F2F2")
ALT_FILL     = PatternFill("solid", "FAFAFA", "FAFAFA")

THIN_BORDER  = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

SPECIAL_REPLACE = {"ccy": "Currency", "amt": "Amount", "pct": "(%)"}

_CAMEL_RE = re.compile(r"[_\s]+")
EXCEL_FORBIDDEN_RE = re.compile(r'[:\\/?*\[\]]')
_SURROGATE_RE = re.compile(r'[\uD800-\uDFFF]')  # unpaired UTF-16 surrogates (invalid in XML)

def _strip_illegal_xml(s: str) -> str:
    """Remove XML-illegal control chars & stray surrogates; preserve normal whitespace."""
    if s is None:
        return ""
    s = str(s)
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    s = _SURROGATE_RE.sub("", s)
    return s


NUMERIC_RE = re.compile(
    r"""
    ^\s*
    (?P<neg>\()?\s*
    (?P<num>
        [-+]?
        (?:
            \d{1,3}(?:,\d{3})+     # 1,234 style
            |\d+                   # plain digits
        )
        (?:\.\d+)?                 # decimal part
    )
    \s*(?P<pct>%|pct)?\s*
    (?(neg)\))\s*$
    """,
    re.VERBOSE | re.IGNORECASE,
)


def _excel_safe_component(s: str) -> str:
    if s is None:
        return "Sheet"
    s = _strip_illegal_xml(str(s))
    s = EXCEL_FORBIDDEN_RE.sub("-", s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = " ".join(s.split())
    s = s.strip().strip("'")
    return s or "Sheet"


def _unique_sheet_name(wb: Workbook, base: str) -> str:
    name = base[:31] or "Sheet"
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        suffix = f" ({i})"
        candidate = (base[: 31 - len(suffix)] + suffix) or "Sheet"
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def camelify(text: str) -> str:
    parts = _CAMEL_RE.split(text.strip("_"))
    return " ".join(SPECIAL_REPLACE.get(p.lower(), p.capitalize()) for p in parts)


def prettify_col(col: str) -> str:
    return " ".join(SPECIAL_REPLACE.get(p.lower(), p.capitalize()) for p in col.split("_"))


# ---- date extraction ----
_DATE_PAT = re.compile(
    r"""
    (?P<long>
        (?:\d{2}[./-]\d{2}[./-]\d{4}) |
        (?:\d{4}[./-]\d{2}[./-]\d{2})
    )
    |
    (?P<compact>\b\d{6}\b)
    """,
    re.VERBOSE,
)


def _parse_compact(s: str) -> str | None:
    if len(s) != 6 or not s.isdigit():
        return None
    d, m, y = int(s[:2]), int(s[2:4]), int(s[4:])
    if 1 <= d <= 31 and 1 <= m <= 12:
        y += 2000 if y < 50 else 1900
        try:
            return datetime(y, m, d).strftime("%Y-%m-%d")
        except ValueError:
            pass
    y, m, d = int(s[:2]), int(s[2:4]), int(s[4:])
    if 1 <= m <= 12 and 1 <= d <= 31:
        y += 2000 if y < 50 else 1900
        try:
            return datetime(y, m, d).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def split_dates(desc: str) -> tuple[str, str]:
    dates: list[str] = []

    def _repl(match):
        full = match.group(0)
        if match.group("long"):
            parts = re.split(r"[./-]", full)
            if len(parts[0]) == 4:
                y, m, d = parts
            else:
                d, m, y = parts
            try:
                dates.append(datetime(int(y), int(m), int(d)).strftime("%Y-%m-%d"))
            except ValueError:
                pass
        else:
            iso = _parse_compact(full)
            if iso:
                dates.append(iso)
        return " "

    cleaned = _DATE_PAT.sub(_repl, desc)
    cleaned = " ".join(cleaned.split())
    dates_str = " - ".join(dates) if dates else "NULL"
    if not cleaned:
        cleaned = " "
    # sanitize results for safety
    return _strip_illegal_xml(cleaned), _strip_illegal_xml(dates_str)


def _coerce_numeric_if_text(v: Any) -> tuple[Any, str | None]:
    """
    Try to coerce strings that look like numbers (even with spaces inside) into numbers.
    Handles %, (negatives), and digit-only strings.
    """
    if not isinstance(v, str):
        return v, None

    # normalize: remove ALL spaces inside the string
    text = re.sub(r"\s+", "", v)
    if not text:
        return v, None

    m = NUMERIC_RE.match(text)
    if not m:
        return v, None

    raw = m.group("num").replace(",", "")
    try:
        num = float(raw) if ("." in raw) else int(raw)
    except ValueError:
        return v, None

    if m.group("neg"):
        num = -float(num)

    if m.group("pct"):
        return float(num) / 100.0, PERCENT_FMT

    return num, None


def _to_excel_scalar(val: Any) -> Any:
    """
    Convert arbitrary Python values to Excel-safe scalars or strings:
    - Keep native types (int/float/bool/datetime), but guard NaN/Inf.
    - Convert bytes -> decoded string (UTF-8/Latin-1 fallback).
    - Convert containers -> JSON string.
    - Strip XML-illegal control characters everywhere.
    """
    if val is None or isinstance(val, (int, float, bool, datetime)):
        if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
            return ""
        return val

    if hasattr(val, "to_pydatetime"):
        try:
            return val.to_pydatetime()
        except Exception:
            return _strip_illegal_xml(str(val))

    if isinstance(val, (bytes, bytearray)):
        try:
            val = val.decode("utf-8", "replace")
        except Exception:
            val = val.decode("latin-1", "replace")
        return _strip_illegal_xml(val)

    if isinstance(val, str):
        return _strip_illegal_xml(val)

    try:
        s = json.dumps(val, ensure_ascii=False, default=str)
    except TypeError:
        s = str(val)
    return _strip_illegal_xml(s)


# ---- dataframe builder ----
def df_from_records(records: Iterable[Any]):
    records = list(records) or [{"Notice": "No rows"}]

    df = pd.DataFrame(records)
    df.rename(columns=lambda c: prettify_col(str(c)), inplace=True)

    if df.columns.duplicated().any():
        counts: dict[str, int] = {}
        new_cols: list[str] = []
        for col in df.columns:
            if col in counts:
                counts[col] += 1
                new_cols.append(f"{col}_{counts[col]}")
            else:
                counts[col] = 0
                new_cols.append(col)
        df.columns = new_cols

    if "Description" in df.columns:
        desc_clean, dates_col = [], []
        for text in df["Description"].astype(str):
            d_clean, d_dates = split_dates(text)
            desc_clean.append(d_clean)
            dates_col.append(d_dates)
        df["Description"] = desc_clean
        desc_idx = df.columns.get_indexer_for(["Description"])[0]
        df.insert(desc_idx + 1, "Dates", dates_col)

    df.fillna("NULL", inplace=True)
    df = df.map(_to_excel_scalar)  # safe types for openpyxl (and sanitized)
    return df


# ---- write helpers ----
def write_df(ws, df: pd.DataFrame, start_row: int = 3):
    pct_header_cols = {i + 1 for i, c in enumerate(df.columns) if "(%)" in c}

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        is_header = (r_idx == start_row)
        for c_idx, value in enumerate(row, 1):
            # attempt numeric coercion if text
            coerced, fmt = _coerce_numeric_if_text(value)
            safe_val = _to_excel_scalar(coerced)

            cell = ws.cell(r_idx, c_idx, safe_val)
            cell.font = HDR_FONT if is_header else DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if is_header:
                cell.fill = HDR_FILL
            else:
                # light zebra striping for readability
                if r_idx % 2 == 0:
                    cell.fill = ALT_FILL

                # apply % format if detected from text OR header
                if fmt == PERCENT_FMT or c_idx in pct_header_cols:
                    cell.number_format = PERCENT_FMT

    # freeze panes under the title row + 1 blank row
    ws.freeze_panes = ws.cell(start_row + 1, 1)

    # auto-size columns with reasonable bounds
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(r, col).value or "") for r in range(1, ws.max_row + 1)]
        width = min(max(10, max(len(v) for v in values) + 2), 60)
        ws.column_dimensions[get_column_letter(col)].width = width


# ---- public API ----
def _ensure_records(seq: Sequence[Any] | None) -> list[dict]:
    seq = seq or []
    out: list[dict] = []
    for x in seq:
        if isinstance(x, dict):
            out.append(x)
        elif isinstance(x, (list, tuple)):
            out.append({f"col_{i+1}": v for i, v in enumerate(x)})
        else:
            out.append({"value": x})
    return out


def json_to_xlsx_bytes(data: Union[dict, str], bank_code: str) -> BytesIO:
    if isinstance(data, str):
        data = json.loads(data)

    wb = Workbook()
    wb.remove(wb.active)

    bank_code = _excel_safe_component(bank_code).replace(" ", "")
    bank_code = (bank_code or "BANK")[:10].upper()

    if not isinstance(data, dict):
        _sheet_from_records(wb, "Root", _ensure_records([data]), bank_code, is_asset=False)
    else:
        for top_key, top_val in data.items():
            is_asset = str(top_key).lower().startswith("asset")

            if isinstance(top_val, list):
                _sheet_from_records(wb, str(top_key), _ensure_records(top_val), bank_code, is_asset)

            elif isinstance(top_val, dict):
                for cls_key, cls_val in top_val.items():
                    if isinstance(cls_val, list):
                        _sheet_from_records(wb, str(cls_key), _ensure_records(cls_val), bank_code, is_asset)
                    elif isinstance(cls_val, dict):
                        for sub_k, sub_v in cls_val.items():
                            if isinstance(sub_v, list):
                                _sheet_from_records(
                                    wb, f"{cls_key}_{sub_k}", _ensure_records(sub_v), bank_code, is_asset
                                )
                            else:
                                _sheet_from_records(
                                    wb, f"{cls_key}_{sub_k}", _ensure_records([sub_v]), bank_code, is_asset
                                )
                    else:
                        _sheet_from_records(wb, str(cls_key), _ensure_records([cls_val]), bank_code, is_asset)

            else:
                _sheet_from_records(wb, str(top_key), _ensure_records([top_val]), bank_code, is_asset)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _sheet_from_records(wb: Workbook, raw_cls: str, recs: list[Any], bank_code: str, is_asset: bool):
    cls_title = camelify(raw_cls)
    base = f"{bank_code}_{_excel_safe_component(cls_title)}".replace(" ", "")
    sheet_name = _unique_sheet_name(wb, base)

    ws = wb.create_sheet(sheet_name)
    title = ws.cell(1, 1, cls_title)
    title.font = TITLE_FONT

    ws.cell(2, 1, "")

    write_df(ws, df_from_records(recs), start_row=3)
