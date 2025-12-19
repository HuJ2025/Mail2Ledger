# app/agents/excel_table_detector.py
from __future__ import annotations
from io import BytesIO
from typing import Any, Dict, List
import json

from openpyxl import load_workbook
from app.utils.excelchecker import ensure_openpyxl_readable_xlsx_bytes
from app.services.openai_service import get_agent_response_bg
from app.utils.helpers import _safe_json_loads

from app.prompts.extract_agent_prompts import EXCEL_HEADER_DATA_RANGE_PROMPT  # 你貼嗰段prompt

def build_workbook_cells_payload(
    xlsx_bytes: bytes,
    max_rows: int = 200,   # ✅ 控制token，先掃頭200行
    max_cols: int = 60,    # ✅ 控制token，先掃頭60列
) -> Dict[str, Any]:
    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    sheets = []

    for ws in wb.worksheets:
        rmax = min(ws.max_row or 1, max_rows)
        cmax = min(ws.max_column or 1, max_cols)

        cells_2d: List[List[Dict[str, Any]]] = []
        for r in range(1, rmax + 1):
            row_list = []
            for c in range(1, cmax + 1):
                v = ws.cell(r, c).value
                # displayed text（簡化：轉字串；空就 ""）
                s = "" if v is None else str(v)
                row_list.append({"r": r, "c": c, "v": s})
            cells_2d.append(row_list)

        sheets.append({"name": ws.title, "cells": cells_2d})

    return {"workbook": {"sheets": sheets}}

def detect_tables_from_bytes(
    content_bytes: bytes,
    password: str | None = None,
    model: str = "gpt-5.2-mini",
) -> Dict[str, Any]:
    xlsx_bytes = ensure_openpyxl_readable_xlsx_bytes(content_bytes, password=password)
    payload = build_workbook_cells_payload(xlsx_bytes)

    resp = get_agent_response_bg(
        user_input=json.dumps(payload, ensure_ascii=False),
        instructions=EXCEL_HEADER_DATA_RANGE_PROMPT,
        model=model,
    )
    return _safe_json_loads(resp)  # { "<sheetName>": [ {header_row,data_start,data_end,table,...}, ...], ...}
